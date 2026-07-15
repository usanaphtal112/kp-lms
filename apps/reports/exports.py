from io import BytesIO

from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font


def build_workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
    )


def autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def export_clinical_reports_workbook(reports):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clinical Reports"

    headers = [
        "Student Registration",
        "Student Name",
        "Module",
        "Title",
        "Facility",
        "Unit",
        "Start Date",
        "End Date",
        "Status",
        "Submitted At",
        "Reviewed By",
        "Reviewed At",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for report in reports:
        profile = getattr(report.student, "student_profile", None)

        worksheet.append(
            [
                getattr(profile, "registration_number", ""),
                report.student.get_full_name(),
                report.module_offering.module.code,
                report.title,
                report.facility_name,
                report.department_or_unit,
                report.clinical_start_date,
                report.clinical_end_date,
                report.get_status_display(),
                report.submitted_at,
                str(report.reviewed_by or ""),
                report.reviewed_at,
            ]
        )

    autosize_columns(worksheet)

    return build_workbook_response(
        workbook,
        "clinical-reports.xlsx",
    )


def export_clinical_teaching_reports_workbook(reports):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clinical Teaching"

    headers = [
        "Lecturer",
        "Module",
        "Title",
        "Facility",
        "Unit",
        "Teaching Date",
        "Topic",
        "Students Supervised",
        "Status",
        "Submitted At",
        "Reviewed By",
        "Reviewed At",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for report in reports:
        worksheet.append(
            [
                report.lecturer.get_full_name() or report.lecturer.username,
                report.module_offering.module.code,
                report.title,
                report.facility_name,
                report.department_or_unit,
                report.teaching_date,
                report.topic_taught,
                report.students_supervised_count,
                report.get_status_display(),
                report.submitted_at,
                str(report.reviewed_by or ""),
                report.reviewed_at,
            ]
        )

    autosize_columns(worksheet)

    return build_workbook_response(
        workbook,
        "clinical-teaching-reports.xlsx",
    )


def export_student_portfolio_workbook(student, summary):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Portfolio Summary"

    profile = getattr(student, "student_profile", None)

    worksheet.append(["Student", student.get_full_name() or student.username])
    worksheet.append(["Registration Number", getattr(profile, "registration_number", "")])
    worksheet.append([])

    worksheet.append(["Section", "Module", "Title", "Status / Mark", "Date"])

    for snapshot in summary["attendance_snapshots"]:
        worksheet.append(
            [
                "Attendance",
                snapshot.module_offering.module.code,
                "Attendance eligibility",
                f"{snapshot.attendance_percentage}%",
                snapshot.calculated_at,
            ]
        )

    for log in summary["procedure_logs"]:
        worksheet.append(
            [
                "Procedure Log",
                log.module_offering.module.code,
                log.procedure.name,
                log.get_status_display(),
                log.verified_at,
            ]
        )

    for result in summary["osce_results"]:
        worksheet.append(
            [
                "OSCE Result",
                result.attempt.osce_exam.module_offering.module.code,
                result.attempt.osce_exam.title,
                f"{result.final_mark}%",
                result.published_at,
            ]
        )

    for report in summary["clinical_reports"]:
        worksheet.append(
            [
                "Clinical Report",
                report.module_offering.module.code,
                report.title,
                report.get_status_display(),
                report.submitted_at,
            ]
        )

    for item in summary["manual_items"]:
        worksheet.append(
            [
                item.get_item_type_display(),
                item.module_offering.module.code if item.module_offering else "",
                item.title,
                "Portfolio item",
                item.created_at,
            ]
        )

    autosize_columns(worksheet)

    return build_workbook_response(
        workbook,
        f"portfolio-{student.username}.xlsx",
    )