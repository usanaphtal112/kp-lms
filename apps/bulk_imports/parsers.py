import csv
import io
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_EXTENSIONS = {".csv", ".xlsx"}


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def normalize_cell(value):
    if value is None:
        return ""

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    return str(value).strip()


def read_tabular_file(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()

    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("Only .csv and .xlsx files are supported.")

    uploaded_file.seek(0)

    if extension == ".csv":
        return read_csv(uploaded_file)

    return read_xlsx(uploaded_file)


def read_csv(uploaded_file):
    content = uploaded_file.read().decode("utf-8-sig")
    csv_file = io.StringIO(content)
    reader = csv.DictReader(csv_file)

    if not reader.fieldnames:
        raise ValueError("The CSV file has no header row.")

    rows = []

    for row in reader:
        normalized_row = {
            normalize_header(key): normalize_cell(value)
            for key, value in row.items()
        }
        rows.append(normalized_row)

    return rows


def read_xlsx(uploaded_file):
    workbook = load_workbook(
        filename=io.BytesIO(uploaded_file.read()),
        read_only=True,
        data_only=True,
    )
    worksheet = workbook.active

    rows_iterator = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iterator)
    except StopIteration as exc:
        raise ValueError("The Excel file is empty.") from exc

    normalized_headers = [normalize_header(header) for header in headers]

    if not any(normalized_headers):
        raise ValueError("The Excel file has no header row.")

    rows = []

    for values in rows_iterator:
        row = {
            normalized_headers[index]: normalize_cell(value)
            for index, value in enumerate(values)
            if index < len(normalized_headers) and normalized_headers[index]
        }

        if any(row.values()):
            rows.append(row)

    return rows